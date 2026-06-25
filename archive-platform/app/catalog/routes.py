import io
import os
import re
import datetime
from flask import render_template, request, jsonify, url_for, send_file, current_app, abort
from flask_login import login_required, current_user
from sqlalchemy import or_
from app.catalog import catalog_bp
from app.models import Archive, OperationLog, SpecialTopicType
from app.extensions import db


@catalog_bp.route("/")
@login_required
def list():
    """档案总库列表页"""
    return render_template("catalog/list.html")


@catalog_bp.route("/api")
@login_required
def api_list():
    """DataTables JSON API"""
    draw = int(request.args.get("draw", 1))
    start = int(request.args.get("start", 0))
    length = int(request.args.get("length", 50))
    search_value = request.args.get("search[value]", "").strip()
    category = request.args.get("category", "").strip()
    archive_year = request.args.get("archive_year", "").strip()
    retention_period = request.args.get("retention_period", "").strip()
    responsible = request.args.get("responsible", "").strip()
    object_type = request.args.get("object_type", "").strip()
    topic_type = request.args.get("topic_type", "").strip()

    query = Archive.query

    if category:
        query = query.filter(Archive.category == category)
    if archive_year:
        try:
            query = query.filter(Archive.archive_year == int(archive_year))
        except ValueError:
            pass
    if retention_period:
        query = query.filter(Archive.retention_period == retention_period)
    if object_type:
        query = query.filter(Archive.object_type == object_type)
    if topic_type:
        query = query.filter(Archive.topic_type == topic_type)
    if responsible:
        query = query.filter(Archive.responsible.ilike(f"%{responsible}%"))
    if search_value:
        pattern = f"%{search_value}%"
        query = query.filter(
            or_(
                Archive.title.ilike(pattern),
                Archive.archive_number.ilike(pattern),
                Archive.responsible.ilike(pattern),
                Archive.doc_number.ilike(pattern),
                Archive.keywords.ilike(pattern),
                Archive.remarks.ilike(pattern),
            )
        )

    total = query.count()

    order_col = int(request.args.get("order[0][column]", 0))
    order_dir = request.args.get("order[0][dir]", "asc")
    col_map = {
        0: Archive.id,
        1: Archive.category,
        2: Archive.archive_number,
        3: Archive.title,
        4: Archive.archive_year,
        5: Archive.retention_period,
        6: Archive.responsible,
        7: Archive.doc_date,
    }
    order_by = col_map.get(order_col, Archive.id)
    order_by = order_by.desc() if order_dir == "desc" else order_by.asc()

    rows = query.order_by(order_by).offset(start).limit(length).all()

    data = []
    for a in rows:
        data.append({
            "id": a.id,
            "category": a.category,
            "object_type": a.object_type or "",
            "topic_type": a.topic_type or "",
            "archive_number": a.archive_number or "",
            "title": a.title[:80] if a.title else "",
            "archive_year": a.archive_year or "",
            "retention_period": a.retention_period or "",
            "responsible": a.responsible or "",
            "doc_date": a.doc_date.strftime("%Y-%m-%d") if a.doc_date else "",
            "has_electronic": bool(a.electronic_path),
            "bc_hash": a.bc_hash[:8] + "..." if a.bc_hash else "",
        })

    return jsonify({
        "draw": draw,
        "recordsTotal": total,
        "recordsFiltered": total,
        "data": data,
    })


@catalog_bp.route("/api/stats")
@login_required
def api_stats():
    """筛选器统计信息"""
    from sqlalchemy import func

    categories = db.session.query(
        Archive.category, func.count(Archive.id)
    ).group_by(Archive.category).all()

    years = db.session.query(
        Archive.archive_year, func.count(Archive.id)
    ).filter(Archive.archive_year.isnot(None)).group_by(
        Archive.archive_year
    ).order_by(Archive.archive_year.desc()).all()

    periods = db.session.query(
        Archive.retention_period, func.count(Archive.id)
    ).filter(Archive.retention_period != "").group_by(
        Archive.retention_period
    ).all()

    return jsonify({
        "categories": [{"name": c, "count": n} for c, n in categories],
        "years": [{"year": y, "count": n} for y, n in years],
        "periods": [{"name": p, "count": n} for p, n in periods],
        "total": Archive.query.count(),
        "with_electronic": Archive.query.filter(Archive.electronic_path != "").count(),
        "with_blockchain": Archive.query.filter(Archive.bc_hash != "").count(),
    })


@catalog_bp.route("/api/responsible")
@login_required
def api_responsible():
    """责任者搜索接口，返回匹配的责任者列表（按数量排序）"""
    from sqlalchemy import func
    q = request.args.get("q", "").strip()
    query = db.session.query(
        Archive.responsible, func.count(Archive.id).label("cnt")
    ).filter(Archive.responsible != None, Archive.responsible != "")
    if q:
        query = query.filter(Archive.responsible.ilike(f"%{q}%"))
    rows = query.group_by(Archive.responsible).order_by(
        func.count(Archive.id).desc()
    ).limit(50).all()
    return jsonify([{"name": r, "count": c} for r, c in rows])


# ─── 专题档案类型管理 ─────────────────────────────────────────────────────────

@catalog_bp.route("/api/topic-types")
@login_required
def api_topic_types():
    """获取所有专题档案类型（含禁用的，管理员可见；普通用户只看启用的）"""
    if current_user.is_admin():
        types = SpecialTopicType.query.order_by(SpecialTopicType.sort_order, SpecialTopicType.id).all()
    else:
        types = SpecialTopicType.query.filter_by(is_active=True).order_by(
            SpecialTopicType.sort_order, SpecialTopicType.id
        ).all()
    return jsonify([t.to_dict() for t in types])


@catalog_bp.route("/api/topic-types", methods=["POST"])
@login_required
def api_topic_types_create():
    """新增专题档案类型（管理员）"""
    if not current_user.is_admin():
        return jsonify({"ok": False, "msg": "仅管理员可操作"}), 403
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "msg": "类型名称不能为空"})
    if SpecialTopicType.query.filter_by(name=name).first():
        return jsonify({"ok": False, "msg": f"类型名称 '{name}' 已存在"})
    max_order = db.session.query(db.func.max(SpecialTopicType.sort_order)).scalar() or 0
    t = SpecialTopicType(
        name=name,
        description=(data.get("description") or "").strip(),
        sort_order=int(data.get("sort_order", max_order + 1)),
        is_active=bool(data.get("is_active", True)),
    )
    db.session.add(t)
    db.session.add(OperationLog(
        user_id=current_user.id, action="create",
        target_type="topic_type", target_id=0,
        detail=f"新增专题类型：{name}", ip_address=request.remote_addr,
    ))
    db.session.commit()
    return jsonify({"ok": True, "id": t.id, "data": t.to_dict()})


@catalog_bp.route("/api/topic-types/<int:tid>", methods=["POST"])
@login_required
def api_topic_types_update(tid):
    """修改专题档案类型（管理员）"""
    if not current_user.is_admin():
        return jsonify({"ok": False, "msg": "仅管理员可操作"}), 403
    t = SpecialTopicType.query.get_or_404(tid)
    data = request.json or {}
    new_name = (data.get("name") or "").strip()
    if new_name and new_name != t.name:
        conflict = SpecialTopicType.query.filter_by(name=new_name).first()
        if conflict and conflict.id != tid:
            return jsonify({"ok": False, "msg": f"类型名称 '{new_name}' 已存在"})
        t.name = new_name
    if "description" in data:
        t.description = (data["description"] or "").strip()
    if "sort_order" in data:
        try:
            t.sort_order = int(data["sort_order"])
        except (TypeError, ValueError):
            pass
    if "is_active" in data:
        t.is_active = bool(data["is_active"])
    db.session.add(OperationLog(
        user_id=current_user.id, action="update",
        target_type="topic_type", target_id=tid,
        detail=f"修改专题类型：{t.name}", ip_address=request.remote_addr,
    ))
    db.session.commit()
    return jsonify({"ok": True, "data": t.to_dict()})


@catalog_bp.route("/api/topic-types/<int:tid>/delete", methods=["POST"])
@login_required
def api_topic_types_delete(tid):
    """删除专题档案类型（管理员；若已有档案使用该类型则禁止删除）"""
    if not current_user.is_admin():
        return jsonify({"ok": False, "msg": "仅管理员可操作"}), 403
    t = SpecialTopicType.query.get_or_404(tid)
    count = Archive.query.filter_by(category="专题", topic_type=t.name).count()
    if count > 0:
        return jsonify({"ok": False, "msg": f"该类型下还有 {count} 条档案，无法删除。请先修改这些档案的类型后再删除。"})
    db.session.add(OperationLog(
        user_id=current_user.id, action="delete",
        target_type="topic_type", target_id=tid,
        detail=f"删除专题类型：{t.name}", ip_address=request.remote_addr,
    ))
    db.session.delete(t)
    db.session.commit()
    return jsonify({"ok": True})


# ─── 批量修改 ────────────────────────────────────────────────────────────────

@catalog_bp.route("/api/batch-update", methods=["POST"])
@login_required
def api_batch_update():
    """批量修改档案字段"""
    if not current_user.can_edit():
        return jsonify({"ok": False, "msg": "无编辑权限"}), 403

    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"ok": False, "msg": "未选择档案"})

    # 允许批量修改的字段白名单
    allowed = {"category", "retention_period", "archive_year", "responsible",
               "fonds_number", "class_number", "electronic_location"}
    updates = {k: v for k, v in data.items() if k in allowed and v != "" and v is not None}

    if not updates:
        return jsonify({"ok": False, "msg": "没有需要修改的字段"})

    try:
        # archive_year 需要转 int
        if "archive_year" in updates:
            try:
                updates["archive_year"] = int(updates["archive_year"])
            except (ValueError, TypeError):
                updates.pop("archive_year")

        affected = Archive.query.filter(Archive.id.in_(ids)).update(
            updates, synchronize_session="fetch"
        )

        db.session.add(OperationLog(
            user_id=current_user.id,
            action="update",
            target_type="archive",
            target_id=0,
            detail=f"批量修改 {affected} 条档案：{', '.join(f'{k}={v}' for k,v in updates.items())}",
            ip_address=request.remote_addr,
        ))
        db.session.commit()
        return jsonify({"ok": True, "affected": affected})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "msg": str(e)})


# ─── 导出数据（Excel）───────────────────────────────────────────────────────

@catalog_bp.route("/export")
@login_required
def export_data():
    """按当前筛选条件导出档案数据到 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl 未安装"}), 500

    category = request.args.get("category", "").strip()
    archive_year = request.args.get("archive_year", "").strip()
    retention_period = request.args.get("retention_period", "").strip()
    responsible_filter = request.args.get("responsible", "").strip()
    object_type_filter = request.args.get("object_type", "").strip()
    topic_type_filter = request.args.get("topic_type", "").strip()
    search_value = request.args.get("q", "").strip()

    query = Archive.query
    if category:
        query = query.filter(Archive.category == category)
    if archive_year:
        try:
            query = query.filter(Archive.archive_year == int(archive_year))
        except ValueError:
            pass
    if retention_period:
        query = query.filter(Archive.retention_period == retention_period)
    if object_type_filter:
        query = query.filter(Archive.object_type == object_type_filter)
    if topic_type_filter:
        query = query.filter(Archive.topic_type == topic_type_filter)
    if responsible_filter:
        query = query.filter(Archive.responsible.ilike(f"%{responsible_filter}%"))
    if search_value:
        pattern = f"%{search_value}%"
        query = query.filter(or_(
            Archive.title.ilike(pattern),
            Archive.archive_number.ilike(pattern),
            Archive.responsible.ilike(pattern),
            Archive.keywords.ilike(pattern),
        ))

    # 限制最多导出 5000 条，避免全量导出超时
    EXPORT_LIMIT = 5000
    total_count = query.count()
    rows = query.order_by(Archive.id.asc()).limit(EXPORT_LIMIT).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "档案目录"

    # 表头（保留样式，仅表头行）
    headers = [
        "序号", "档案类别", "全宗号", "档号", "分类号", "案卷号",
        "文件题名", "责任者", "文号", "形成日期", "归档年度",
        "保管期限", "页数", "关键词", "备注",
        "电子版路径", "电子版大小", "电子版位置",
        "区块链哈希(SHA-256)", "区块链MD5",
    ]
    header_fill = PatternFill("solid", fgColor="1A365D")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.row_dimensions[1].height = 30

    # 数据行（只设字体，去掉边框和隔行背景，大幅提速）
    normal_font = Font(name="微软雅黑", size=9)
    for row_idx, a in enumerate(rows, 2):
        row_data = [
            row_idx - 1,
            a.category or "",
            a.fonds_number or "",
            a.archive_number or "",
            a.class_number or "",
            a.file_number or "",
            a.title or "",
            a.responsible or "",
            a.doc_number or "",
            a.doc_date.strftime("%Y-%m-%d") if a.doc_date else "",
            a.archive_year or "",
            a.retention_period or "",
            a.pages or "",
            a.keywords or "",
            a.remarks or "",
            a.electronic_path or "",
            a.electronic_size or "",
            a.electronic_location or "",
            a.bc_hash or "",
            a.bc_md5 or "",
        ]
        ws.append(row_data)

    # 列宽
    col_widths = [6, 8, 8, 18, 8, 8, 45, 12, 18, 12, 8, 8, 6, 20, 20, 30, 12, 20, 40, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 若实际数据超出限制，写入提示行
    if total_count > EXPORT_LIMIT:
        tip_row = len(rows) + 2
        ws.cell(row=tip_row, column=1,
                value=f"[提示] 共 {total_count} 条，本次仅导出前 {EXPORT_LIMIT} 条。请使用筛选条件缩小范围后再导出。")
        ws.cell(row=tip_row, column=1).font = Font(color="FF0000", bold=True)

    # 输出到内存
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    label = f"_{category}" if category else ""
    filename = f"档案目录{label}_{now_str}.xlsx"

    db.session.add(OperationLog(
        user_id=current_user.id,
        action="export",
        target_type="archive",
        target_id=0,
        detail=f"导出档案数据 {len(rows)} 条（类别:{category or '全部'}）",
        ip_address=request.remote_addr,
    ))
    db.session.commit()

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── 导出模板 ────────────────────────────────────────────────────────────────

@catalog_bp.route("/export-template")
@login_required
def export_template():
    """下载档案目录导入模板（含示例行）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl 未安装"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "档案目录"

    headers = [
        ("序号", "填写行号，可留空"),
        ("档案类别", "必填：文书 / 基建 / 科研 / 设备 / 会计"),
        ("全宗号", "如：DLYFY"),
        ("档号", "完整档号，如：DLYFY-WS-2023-001"),
        ("分类号", "档案分类号"),
        ("案卷号", "数字"),
        ("文件题名", "必填：文件全称"),
        ("责任者", "形成单位或个人"),
        ("文号", "发文字号，如：卫院字[2023]1号"),
        ("形成日期", "格式：YYYY-MM-DD，如：2023-01-15"),
        ("归档年度", "4位数字，如：2023"),
        ("保管期限", "永久 / 30年 / 10年 / 长期 / 短期"),
        ("页数", "数字"),
        ("关键词", "多个关键词用分号分隔"),
        ("备注", "其他说明"),
    ]

    header_fill = PatternFill("solid", fgColor="1A365D")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    hint_fill = PatternFill("solid", fgColor="E8EDF2")
    hint_font = Font(name="微软雅黑", italic=True, color="6B7280", size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 第1行：字段名
    for col_idx, (h, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28

    # 第2行：说明行
    for col_idx, (_, hint) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.fill = hint_fill
        cell.font = hint_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 24

    # 第3行：示例数据
    example = [
        1, "文书", "DLYFY", "DLYFY-WS-2023-001", "WS", 1,
        "关于印发2023年度工作计划的通知", "院办公室",
        "卫院字[2023]1号", "2023-01-15", 2023,
        "永久", 8, "工作计划;年度计划", "正本",
    ]
    example_fill = PatternFill("solid", fgColor="FFFDE7")
    example_font = Font(name="微软雅黑", size=9, color="374151")
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.fill = example_fill
        cell.font = example_font
        cell.border = border
        if col_idx in (1, 5, 6, 11, 12, 13):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

    ws.row_dimensions[3].height = 22

    # 列宽
    col_widths = [6, 10, 8, 22, 8, 8, 50, 14, 20, 13, 9, 8, 6, 24, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结前两行
    ws.freeze_panes = "A3"

    # 添加一个说明sheet
    ws2 = wb.create_sheet("填写说明")
    notes = [
        ["字段", "说明", "示例"],
        ["档案类别", "必须是以下之一：文书、基建、科研、设备、会计", "文书"],
        ["全宗号", "本院全宗号，统一填写 DLYFY", "DLYFY"],
        ["档号", "唯一标识，格式：全宗号-类别缩写-年度-流水号", "DLYFY-WS-2023-001"],
        ["形成日期", "必须使用 YYYY-MM-DD 格式", "2023-01-15"],
        ["归档年度", "4位数字年份", "2023"],
        ["保管期限", "只能填写：永久、30年、10年、长期、短期", "永久"],
        ["页数", "纯数字，无单位", "8"],
        ["关键词", "多个关键词请用中文分号（；）分隔", "工作计划；年度计划"],
        ["", "", ""],
        ["注意事项", "", ""],
        ["1. 请勿修改第1、2行（表头和说明行），从第3行开始填写数据", "", ""],
        ["2. 档号字段建议确保唯一，重复档号+类别将被系统跳过（增量导入）", "", ""],
        ["3. 导入时系统自动计算 SHA-256 和 MD5 哈希值，无需手动填写", "", ""],
        ["4. 如有多个sheet，系统只解析名称中含【档案目录】的sheet", "", ""],
    ]
    ws2_font = Font(name="微软雅黑", size=10)
    ws2_header_fill = PatternFill("solid", fgColor="1A365D")
    ws2_header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    for r_idx, row in enumerate(notes, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.fill = ws2_header_fill
                cell.font = ws2_header_font
            elif r_idx == 11:
                cell.font = Font(name="微软雅黑", bold=True, size=10)
            else:
                cell.font = ws2_font
            cell.border = Border(left=Side(style="thin", color="DDDDDD"),
                                 right=Side(style="thin", color="DDDDDD"),
                                 top=Side(style="thin", color="DDDDDD"),
                                 bottom=Side(style="thin", color="DDDDDD"))
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "档案目录导入模板.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── 档案详情 ────────────────────────────────────────────────────────────────

@catalog_bp.route("/<int:archive_id>")
@login_required
def detail(archive_id):
    """档案详情页"""
    archive = Archive.query.get_or_404(archive_id)

    transfers = archive.transfers.order_by(
        db.desc("transfer_date")
    ).limit(10).all()
    borrows = archive.borrows.order_by(
        db.desc("borrow_date")
    ).limit(10).all()
    destructions = archive.destructions.order_by(
        db.desc("appraisal_date")
    ).limit(10).all()

    return render_template(
        "catalog/detail.html",
        archive=archive,
        transfers=transfers,
        borrows=borrows,
        destructions=destructions,
    )


@catalog_bp.route("/<int:archive_id>/download")
@login_required
def download_electronic(archive_id):
    """下载档案电子版（仅管理员或已审批借阅用户）"""
    archive = Archive.query.get_or_404(archive_id)
    if not archive.electronic_path:
        abort(404)
    # 权限：管理员可直接下载；普通用户需有已通过的借阅记录
    if current_user.role != "admin":
        from app.models import Borrow
        approved = Borrow.query.filter_by(
            archive_id=archive_id,
            borrower_id=current_user.id,
            status="已通过",
        ).first()
        if not approved:
            abort(403)

    static_folder = current_app.static_folder
    file_path = os.path.join(static_folder, archive.electronic_path)
    if not os.path.isfile(file_path):
        abort(404)

    # 原始文件名：用标题+扩展名
    ext = os.path.splitext(archive.electronic_path)[1]
    safe_title = re.sub(r'[\\/:*?"<>|]', "_", archive.title or "档案")[:80]
    download_name = f"{safe_title}{ext}"

    log = OperationLog(
        user_id=current_user.id,
        action="download",
        target_type="archive",
        target_id=archive_id,
        detail=f"下载电子版：{archive.title}",
        ip_address=request.remote_addr,
    )
    db.session.add(log)
    db.session.commit()

    return send_file(file_path, as_attachment=True, download_name=download_name)


@catalog_bp.route("/<int:archive_id>/preview")
@login_required
def preview_electronic(archive_id):
    """在线预览电子版（PDF/图片直接返回，其他提示下载）"""
    archive = Archive.query.get_or_404(archive_id)
    if not archive.electronic_path:
        abort(404)
    if current_user.role != "admin":
        from app.models import Borrow
        approved = Borrow.query.filter_by(
            archive_id=archive_id,
            borrower_id=current_user.id,
            status="已通过",
        ).first()
        if not approved:
            abort(403)

    static_folder = current_app.static_folder
    file_path = os.path.join(static_folder, archive.electronic_path)
    if not os.path.isfile(file_path):
        abort(404)

    ext = os.path.splitext(archive.electronic_path)[1].lower()
    mime_map = {
        ".pdf": "application/pdf",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".gif": "image/gif",
    }
    mimetype = mime_map.get(ext)
    if mimetype:
        return send_file(file_path, mimetype=mimetype)
    else:
        # 非预览格式，走下载
        return preview_electronic.__wrapped__(archive_id) if hasattr(preview_electronic, '__wrapped__') else download_electronic(archive_id)


@catalog_bp.route("/<int:archive_id>/edit", methods=["GET", "POST"])
@login_required
def edit(archive_id):
    """编辑档案"""
    if not current_user.can_edit():
        return jsonify({"error": "无编辑权限"}), 403

    archive = Archive.query.get_or_404(archive_id)

    if request.method == "POST":
        data = request.form
        archive.category = data.get("category", archive.category)
        archive.archive_number = data.get("archive_number", archive.archive_number)
        archive.fonds_number = data.get("fonds_number", archive.fonds_number)
        try:
            archive.archive_year = int(data["archive_year"]) if data.get("archive_year") else archive.archive_year
        except (ValueError, KeyError):
            pass
        archive.retention_period = data.get("retention_period", archive.retention_period)
        archive.class_number = data.get("class_number", archive.class_number)
        try:
            archive.file_number = int(data["file_number"]) if data.get("file_number") else archive.file_number
        except (ValueError, KeyError):
            pass
        archive.title = data.get("title", archive.title)
        archive.responsible = data.get("responsible", archive.responsible)
        archive.doc_number = data.get("doc_number", archive.doc_number)
        if data.get("doc_date"):
            try:
                archive.doc_date = datetime.date.fromisoformat(data["doc_date"])
            except ValueError:
                pass
        archive.pages = data.get("pages", archive.pages)
        archive.keywords = data.get("keywords", archive.keywords)
        archive.remarks = data.get("remarks", archive.remarks)
        archive.electronic_location = data.get("electronic_location", archive.electronic_location)

        log = OperationLog(
            user_id=current_user.id,
            action="update",
            target_type="archive",
            target_id=archive.id,
            detail=f"编辑档案: {archive.title[:50] if archive.title else archive.archive_number}",
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({"success": True, "message": "保存成功"})

    return render_template("catalog/edit.html", archive=archive)
