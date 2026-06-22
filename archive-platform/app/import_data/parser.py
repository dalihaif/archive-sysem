"""
Excel 档案数据解析器
支持文书/设备/科研/基建/会计五类档案的批量导入
自动字段映射、档号去重、区块链哈希计算
"""

import datetime
import hashlib
import os
import re

import openpyxl

# 字段映射：Excel列名 → 模型字段
FIELD_MAP = {
    "序号": "seq",
    "档号": "archive_number",
    "全宗号": "fonds_number",
    "归档年度": "archive_year",
    "保管期限": "retention_period",
    "分类号": "class_number",
    "件号": "file_number",
    "文件题名": "title",
    "责任者": "responsible",
    "文件编号": "doc_number",
    "日期": "doc_date",
    "形成日期": "doc_date",
    "页数": "pages",
    "页号": "pages",
    "主题词": "keywords",
    "备注": "remarks",
    "附注": "remarks",
}

# 类别识别（从文件名推断）
CATEGORY_MAP = {
    "文书": "文书",
    "基建": "基建",
    "科研": "科研",
    "设备": "设备",
    "科技": "设备",
    "会计": "会计",
}


def detect_category(filename: str) -> str:
    """从文件名推断档案类别"""
    for key, value in CATEGORY_MAP.items():
        if key in filename:
            return value
    return "文书"


def map_row_to_dict(headers: list, row: tuple) -> dict:
    """将 Excel 行数据映射为字典"""
    data = {}
    for i, header in enumerate(headers):
        if header is None:
            continue
        header_str = str(header).strip()
        if header_str not in FIELD_MAP:
            continue
        field = FIELD_MAP[header_str]
        value = row[i] if i < len(row) else None

        if value is None:
            continue

        # 类型转换
        if field == "archive_year":
            try:
                data[field] = int(float(value))
            except (ValueError, TypeError):
                data[field] = None
        elif field == "file_number":
            try:
                data[field] = int(float(value))
            except (ValueError, TypeError):
                data[field] = None
        elif field == "doc_date":
            data[field] = parse_date(value)
        elif field == "seq":
            # 序号不导入，跳过
            continue
        else:
            val_str = str(value).strip()
            if val_str and val_str.lower() not in ("none", "null", "nan"):
                data[field] = val_str

    return data


def parse_date(value) -> datetime.date | None:
    """解析各种日期格式"""
    if value is None:
        return None

    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    # 字符串格式
    val_str = str(value).strip()
    if not val_str or val_str.lower() in ("none", "null", "nan"):
        return None

    # 尝试多种格式
    patterns = [
        (r"^(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})[日]?$", "%Y-%m-%d"),
        (r"^(\d{4})(\d{2})(\d{2})$", "%Y%m%d"),
        (r"^(\d{4})$", "%Y"),
    ]

    for pattern, fmt in patterns:
        m = re.match(pattern, val_str)
        if m:
            groups = m.groups()
            if len(groups) == 3:
                return datetime.date(int(groups[0]), int(groups[1]), int(groups[2]))
            elif len(groups) == 1:
                return datetime.date(int(groups[0]), 1, 1)

    return None


def compute_hashes_from_row(data: dict) -> tuple:
    """从字典数据计算 SHA-256 和 MD5"""
    content = ""
    for key in sorted(data.keys()):
        val = data.get(key)
        if val:
            content += f"{key}={val};"
    bc_hash = hashlib.sha256(content.encode("utf-8")).hexdigest()
    bc_md5 = hashlib.md5(content.encode("utf-8")).hexdigest()
    return bc_hash, bc_md5


def parse_excel(filepath: str) -> list[dict]:
    """解析 Excel 文件，返回档案数据列表（只解析档案目录sheet）"""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")

    filename = os.path.basename(filepath)
    category = detect_category(filename)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results = []

    # 找到档案目录sheet（兼容尾部空格）
    target_sheet = None
    for sn in wb.sheetnames:
        if "档案目录" in sn.strip():
            target_sheet = sn
            break

    if target_sheet is None:
        wb.close()
        return results

    ws = wb[target_sheet]

    # 找表头行：前4行中查找含"档号"的行
    header_row = None
    headers = []
    rows_cache = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True), 1):
        rows_cache.append(row)
        if any("档号" in (str(v or "")) for v in row):
            header_row = i
            headers = row
            break

    if header_row is None:
        wb.close()
        return results

    # 从表头下一行开始读取数据
    row_count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # 跳过全空行
        if all(v is None for v in row):
            continue
        data = map_row_to_dict(headers, row)
        if data.get("archive_number") or data.get("title"):
            data["_category"] = category
            bc_hash, bc_md5 = compute_hashes_from_row(data)
            data["bc_hash"] = bc_hash
            data["bc_md5"] = bc_md5
            results.append(data)
        row_count += 1

    wb.close()
    return results


def import_data_to_db(filepath: str, app, batch_name: str = None) -> dict:
    """将解析结果导入数据库（增量：按档号去重，批量高效导入）"""
    from app.models import Archive
    from app.extensions import db

    records = parse_excel(filepath)
    if not records:
        return {"imported": 0, "skipped": 0, "total": 0, "message": "未解析到有效数据"}

    if batch_name is None:
        batch_name = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    # 批量加载已有档号（高效去重，避免N+1查询）
    existing_keys = set()
    rows = db.session.query(Archive.archive_number, Archive.category).all()
    for an, cat in rows:
        existing_keys.add((an, cat))

    imported = 0
    skipped = 0
    batch = []

    for data in records:
        archive_number = data.get("archive_number", "")
        category = data.pop("_category", "文书")

        key = (archive_number, category)
        if key in existing_keys:
            skipped += 1
            continue

        archive = Archive(
            category=category,
            archive_number=archive_number,
            fonds_number=data.get("fonds_number", ""),
            archive_year=data.get("archive_year"),
            retention_period=data.get("retention_period", ""),
            class_number=data.get("class_number", ""),
            file_number=data.get("file_number"),
            title=data.get("title", ""),
            responsible=data.get("responsible", ""),
            doc_number=data.get("doc_number", ""),
            doc_date=data.get("doc_date"),
            pages=str(data.get("pages", "")) if data.get("pages") else "",
            keywords=data.get("keywords", ""),
            remarks=data.get("remarks", ""),
            bc_hash=data.get("bc_hash", ""),
            bc_md5=data.get("bc_md5", ""),
            import_batch=batch_name,
        )
        batch.append(archive)
        imported += 1

        # 每1000条批量插入
        if len(batch) >= 1000:
            db.session.add_all(batch)
            db.session.commit()
            batch = []

    # 剩余
    if batch:
        db.session.add_all(batch)
        db.session.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "total": len(records),
        "message": f"导入 {imported} 条，跳过 {skipped} 条（已存在），共 {len(records)} 条",
        "batch": batch_name,
        "category": records[0].get("_category", category) if records else "未知",
    }
