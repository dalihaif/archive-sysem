import secrets
import datetime
import io
from flask import render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from sqlalchemy import or_
from app.borrow import borrow_bp
from app.models import Borrow, Archive, OperationLog
from app.extensions import db


# ─────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────
def _log(action, target_id, detail):
    db.session.add(OperationLog(
        user_id=current_user.id,
        action=action,
        target_type="borrow",
        target_id=target_id,
        detail=detail,
        ip_address=request.remote_addr,
    ))


# ─────────────────────────────────────────────
# 借阅管理主页
# ─────────────────────────────────────────────
@borrow_bp.route("/")
@login_required
def list():
    pending_count = Borrow.query.filter_by(status="待审批").count()
    overdue_count = Borrow.query.filter(
        Borrow.status == "已通过",
        Borrow.return_date < datetime.date.today()
    ).count()
    return render_template(
        "borrow/list.html",
        pending_count=pending_count,
        overdue_count=overdue_count,
    )


# ─────────────────────────────────────────────
# API: 借阅列表（DataTables 服务端分页）
# ─────────────────────────────────────────────
@borrow_bp.route("/api/list")
@login_required
def api_list():
    draw   = request.args.get("draw", 1, type=int)
    start  = request.args.get("start", 0, type=int)
    length = request.args.get("length", 20, type=int)
    search = request.args.get("search[value]", "").strip()
    status = request.args.get("status", "")

    q = Borrow.query

    if status:
        q = q.filter(Borrow.status == status)
    if search:
        q = q.filter(or_(
            Borrow.borrower.contains(search),
            Borrow.borrower_department.contains(search),
            Borrow.archive_ref.contains(search),
        ))

    total = q.count()
    rows  = q.order_by(Borrow.created_at.desc()).offset(start).limit(length).all()

    data = []
    today = datetime.date.today()
    for b in rows:
        is_overdue = (b.status == "已通过" and b.return_date and b.return_date < today)
        status_badge = {
            "待审批": '<span class="badge bg-warning text-dark">待审批</span>',
            "已通过": '<span class="badge bg-success">已通过</span>' if not is_overdue else '<span class="badge bg-danger">已逾期</span>',
            "已拒绝": '<span class="badge bg-secondary">已拒绝</span>',
            "已归还": '<span class="badge bg-info">已归还</span>',
        }.get(b.status, b.status)

        data.append({
            "id": b.id,
            "borrower": b.borrower or "",
            "borrower_department": b.borrower_department or "",
            "archive_ref": (b.archive_ref or "")[:40],
            "purpose": (b.purpose or "")[:30],
            "borrow_date": b.borrow_date.strftime("%Y-%m-%d") if b.borrow_date else "",
            "return_date": b.return_date.strftime("%Y-%m-%d") if b.return_date else "",
            "status": status_badge,
            "status_raw": b.status,
            "created_at": b.created_at.strftime("%Y-%m-%d %H:%M") if b.created_at else "",
        })

    return jsonify({
        "draw": draw,
        "recordsTotal": total,
        "recordsFiltered": total,
        "data": data,
    })


# ─────────────────────────────────────────────
# API: 借阅详情
# ─────────────────────────────────────────────
@borrow_bp.route("/api/<int:bid>")
@login_required
def api_detail(bid):
    b = Borrow.query.get_or_404(bid)
    archive = None
    if b.archive_id:
        a = Archive.query.get(b.archive_id)
        if a:
            archive = {"title": a.title, "archive_number": a.archive_number, "category": a.category}
    return jsonify({
        "id": b.id,
        "borrower": b.borrower,
        "borrower_department": b.borrower_department or "",
        "borrower_phone": b.borrower_phone or "",
        "archive_ref": b.archive_ref or "",
        "purpose": b.purpose or "",
        "borrow_date": b.borrow_date.strftime("%Y-%m-%d") if b.borrow_date else "",
        "return_date": b.return_date.strftime("%Y-%m-%d") if b.return_date else "",
        "status": b.status,
        "approve_comment": b.approve_comment or "",
        "approve_time": b.approve_time.strftime("%Y-%m-%d %H:%M") if b.approve_time else "",
        "approver": b.approver.real_name or b.approver.username if b.approver else "",
        "access_code": b.access_code or "",
        "archive": archive,
        "can_view_electronic": b.can_view_electronic(),
        "created_at": b.created_at.strftime("%Y-%m-%d %H:%M") if b.created_at else "",
    })


# ─────────────────────────────────────────────
# API: 审批（通过/拒绝）
# ─────────────────────────────────────────────
@borrow_bp.route("/api/<int:bid>/approve", methods=["POST"])
@login_required
def api_approve(bid):
    if not current_user.can_edit():
        return jsonify({"ok": False, "msg": "权限不足"}), 403

    b = Borrow.query.get_or_404(bid)
    if b.status != "待审批":
        return jsonify({"ok": False, "msg": "该申请已审批"})

    action = request.json.get("action")   # "approve" | "reject"
    comment = request.json.get("comment", "")

    if action == "approve":
        b.status = "已通过"
        b.access_code = secrets.token_hex(3).upper()   # 6 位验证码
        detail = f"审批通过借阅申请#{bid}（{b.borrower}）"
    elif action == "reject":
        b.status = "已拒绝"
        detail = f"拒绝借阅申请#{bid}（{b.borrower}）"
    else:
        return jsonify({"ok": False, "msg": "非法操作"})

    b.approve_comment = comment
    b.approver_id = current_user.id
    b.approve_time = datetime.datetime.utcnow()
    _log("borrow", bid, detail)
    db.session.commit()
    return jsonify({"ok": True, "access_code": b.access_code or ""})


# ─────────────────────────────────────────────
# API: 归还登记
# ─────────────────────────────────────────────
@borrow_bp.route("/api/<int:bid>/return", methods=["POST"])
@login_required
def api_return(bid):
    if not current_user.can_edit():
        return jsonify({"ok": False, "msg": "权限不足"}), 403

    b = Borrow.query.get_or_404(bid)
    if b.status not in ("已通过",):
        return jsonify({"ok": False, "msg": "当前状态不可归还"})

    b.status = "已归还"
    b.return_date = datetime.date.today()
    _log("borrow", bid, f"确认归还借阅#{bid}（{b.borrower}）")
    db.session.commit()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────
# API: 新建借阅申请（PC端管理员代录）
# ─────────────────────────────────────────────
@borrow_bp.route("/api/create", methods=["POST"])
@login_required
def api_create():
    if not current_user.can_edit():
        return jsonify({"ok": False, "msg": "权限不足"}), 403

    data = request.json or {}
    b = Borrow(
        borrower=data.get("borrower", "").strip(),
        borrower_department=data.get("borrower_department", "").strip(),
        borrower_phone=data.get("borrower_phone", "").strip(),
        archive_ref=data.get("archive_ref", "").strip(),
        purpose=data.get("purpose", "").strip(),
        borrow_date=datetime.date.today(),
        status="待审批",
    )
    if not b.borrower:
        return jsonify({"ok": False, "msg": "借阅人不能为空"})

    db.session.add(b)
    db.session.flush()
    _log("borrow", b.id, f"录入借阅申请（{b.borrower}）")
    db.session.commit()
    return jsonify({"ok": True, "id": b.id})


# ─────────────────────────────────────────────
# API: 修改借阅记录
# ─────────────────────────────────────────────
@borrow_bp.route("/api/<int:bid>/update", methods=["POST"])
@login_required
def api_update(bid):
    if not current_user.can_edit():
        return jsonify({"ok": False, "msg": "权限不足"}), 403

    b = Borrow.query.get_or_404(bid)
    data = request.json or {}

    changes = []
    field_map = {
        "borrower": ("borrower", str),
        "borrower_department": ("borrower_department", str),
        "borrower_phone": ("borrower_phone", str),
        "archive_ref": ("archive_ref", str),
        "purpose": ("purpose", str),
    }

    for json_key, (model_key, _) in field_map.items():
        if json_key in data:
            new_val = (data[json_key] or "").strip()
            old_val = getattr(b, model_key) or ""
            if new_val != old_val:
                changes.append(f"{json_key}: {old_val!r} -> {new_val!r}")
                setattr(b, model_key, new_val)

    # 归还日期（特殊处理）
    if "return_date" in data:
        rd_str = (data["return_date"] or "").strip()
        if rd_str:
            try:
                rd = datetime.datetime.strptime(rd_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"ok": False, "msg": "归还日期格式错误，需 YYYY-MM-DD"})
            if b.return_date != rd:
                changes.append(f"return_date: {b.return_date} -> {rd}")
                b.return_date = rd
        else:
            if b.return_date is not None:
                changes.append("return_date: cleared")
                b.return_date = None

    # 状态（仅管理员可改状态）
    if "status" in data and current_user.is_admin():
        new_status = (data["status"] or "").strip()
        valid_statuses = ["待审批", "已通过", "已拒绝", "已归还"]
        if new_status in valid_statuses and b.status != new_status:
            changes.append(f"status: {b.status} -> {new_status}")
            b.status = new_status
            # 如果改为已通过且没有 access_code，自动生成
            if new_status == "已通过" and not b.access_code:
                b.access_code = secrets.token_hex(3).upper()
                b.approver_id = current_user.id
                b.approve_time = datetime.datetime.utcnow()

    if not changes:
        return jsonify({"ok": False, "msg": "没有修改任何字段"})

    detail = f"修改借阅记录#{bid}（{b.borrower}）：{'; '.join(changes)}"
    _log("borrow", bid, detail)
    db.session.commit()
    return jsonify({"ok": True, "changes": len(changes)})


# ─────────────────────────────────────────────
# API: 导出借阅数据（Excel）
# ─────────────────────────────────────────────
@borrow_bp.route("/export")
@login_required
def export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    status = request.args.get("status", "").strip()
    search = request.args.get("search", "").strip()

    q = Borrow.query
    if status:
        q = q.filter(Borrow.status == status)
    if search:
        q = q.filter(or_(
            Borrow.borrower.contains(search),
            Borrow.borrower_department.contains(search),
            Borrow.archive_ref.contains(search),
        ))

    rows = q.order_by(Borrow.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "借阅记录"

    headers = ["ID", "借阅人", "所在部门", "联系电话", "借阅档案", "借阅目的",
               "登记日期", "归还日期", "状态", "审批人", "审批意见", "访问验证码", "创建时间"]

    # 表头样式
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    today = datetime.date.today()
    for row_idx, b in enumerate(rows, 2):
        is_overdue = (b.status == "已通过" and b.return_date and b.return_date < today)
        display_status = "已逾期" if is_overdue else b.status

        values = [
            b.id,
            b.borrower or "",
            b.borrower_department or "",
            b.borrower_phone or "",
            b.archive_ref or "",
            b.purpose or "",
            b.borrow_date.strftime("%Y-%m-%d") if b.borrow_date else "",
            b.return_date.strftime("%Y-%m-%d") if b.return_date else "",
            display_status,
            b.approver.real_name or b.approver.username if b.approver else "",
            b.approve_comment or "",
            b.access_code or "",
            b.created_at.strftime("%Y-%m-%d %H:%M") if b.created_at else "",
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            # 逾期行标红
            if is_overdue:
                cell.fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")

    # 列宽
    col_widths = [6, 12, 18, 14, 30, 30, 12, 12, 10, 10, 24, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
    # 修正列宽设置（支持超过26列）
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"borrows_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ─────────────────────────────────────────────
# API: 删除
# ─────────────────────────────────────────────
@borrow_bp.route("/api/<int:bid>/delete", methods=["POST"])
@login_required
def api_delete(bid):
    if not current_user.is_admin():
        return jsonify({"ok": False, "msg": "需要管理员权限"}), 403
    b = Borrow.query.get_or_404(bid)
    _log("delete", bid, f"删除借阅记录#{bid}（{b.borrower}）")
    db.session.delete(b)
    db.session.commit()
    return jsonify({"ok": True})
