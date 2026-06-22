"""
导出会计档案跳过记录列表
- 找出每个Sheet中档号重复的记录（保留第1条，后续的都是"跳过"记录）
- 导出为Excel文件，含重复情况说明
"""
import sys
import os
import collections
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))

XLSX_PATH = r"E:\工作文档\档案室\02_归档文件目录与总库\档案总库\会计档案.xlsx"
OUT_PATH   = r"E:\工作文档\档案室\02_归档文件目录与总库\档案总库\会计档案_跳过记录.xlsx"

SHEET_CATEGORY = {
    "会计凭证": "会计凭证",
    "会计报表": "会计报表",
    "会计账簿": "会计账簿",
    "报表-20":  "报表-20",
}

def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def main():
    wb_in  = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)   # 删除默认Sheet

    total_skipped = 0

    for sn in ["会计凭证", "会计报表", "会计账簿", "报表-20"]:
        if sn not in wb_in.sheetnames:
            print(f"[警告] Sheet [{sn}] 不存在，跳过")
            continue

        ws_in  = wb_in[sn]
        rows   = list(ws_in.iter_rows(values_only=True))
        header = [str(h).strip() if h is not None else "" for h in rows[0]]

        # 找件号列
        fn_idx = next((i for i, h in enumerate(header) if h == "件号"), None)
        if fn_idx is None:
            print(f"[警告] Sheet [{sn}] 未找到件号列，跳过")
            continue

        # 计算档号：Sheet名 + "-" + 件号
        seen    = {}          # archive_number -> 首次出现的行号(1-based, 含header=行1)
        skipped = []          # (原始Excel行号, row_data, archive_number, dup_count)

        for row_idx, row in enumerate(rows[1:], start=2):   # row_idx 从 2 开始（含header行）
            if all(v is None for v in row):
                continue
            fn_val = row[fn_idx] if fn_idx < len(row) else None
            if fn_val is None:
                continue
            an = f"{sn}-{str(fn_val).strip()}"

            if an not in seen:
                seen[an] = row_idx
            else:
                skipped.append({
                    "sheet":          sn,
                    "excel_row":      row_idx,
                    "archive_number": an,
                    "first_row":      seen[an],
                    "row_data":       row,
                })

        print(f"[{sn}] 跳过 {len(skipped)} 条")
        total_skipped += len(skipped)

        if not skipped:
            continue

        # ── 创建输出Sheet ──
        ws_out = wb_out.create_sheet(title=sn)

        # 表头颜色
        hdr_fill = PatternFill("solid", fgColor="1A365D")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        dup_fill  = PatternFill("solid", fgColor="FFF3CD")   # 淡黄：被跳过行
        border    = thin_border()

        # 写说明行
        ws_out.append([f"【{sn}】被跳过记录（共 {len(skipped)} 条）——原始数据中与首次出现记录档号完全相同"])
        ws_out.merge_cells(start_row=1, start_column=1,
                           end_row=1,   end_column=len(header) + 3)
        ws_out.cell(1, 1).font      = Font(bold=True, color="856404", size=11)
        ws_out.cell(1, 1).fill      = PatternFill("solid", fgColor="FFF3CD")
        ws_out.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws_out.row_dimensions[1].height = 22

        # 写列头：原始列 + 附加说明列
        out_header = header + ["原始Excel行号", "档号（生成）", "首次出现Excel行号"]
        ws_out.append(out_header)
        for col_idx, val in enumerate(out_header, start=1):
            cell = ws_out.cell(2, col_idx, val)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws_out.row_dimensions[2].height = 28

        # 写数据行
        for row_num, rec in enumerate(skipped, start=3):
            row_vals = list(rec["row_data"]) + [
                rec["excel_row"],
                rec["archive_number"],
                rec["first_row"],
            ]
            ws_out.append(row_vals)
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws_out.cell(row_num, col_idx, val)
                cell.fill      = dup_fill
                cell.font      = Font(size=9)
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

        # 列宽自适应（简单版）
        for col_idx, col_name in enumerate(out_header, start=1):
            max_len = max(len(str(col_name)), 8)
            for rec in skipped[:50]:   # 取前50行估算
                v = rec["row_data"][col_idx - 1] if col_idx - 1 < len(rec["row_data"]) else ""
                max_len = max(max_len, len(str(v or "")))
            ws_out.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        # 冻结前两行
        ws_out.freeze_panes = "A3"

    wb_in.close()

    # ── 汇总Sheet ──
    ws_sum = wb_out.create_sheet(title="汇总", index=0)
    ws_sum.append(["会计档案跳过记录汇总"])
    ws_sum.merge_cells("A1:C1")
    ws_sum.cell(1, 1).font      = Font(bold=True, size=14, color="1A365D")
    ws_sum.cell(1, 1).alignment = Alignment(horizontal="center")
    ws_sum.row_dimensions[1].height = 30

    ws_sum.append(["生成时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
    ws_sum.append([""])
    ws_sum.append(["Sheet", "跳过条数", "说明"])
    for cell in ws_sum[4]:
        cell.fill   = PatternFill("solid", fgColor="1A365D")
        cell.font   = Font(bold=True, color="FFFFFF")
        cell.border = thin_border()

    sheet_stats = {"会计凭证": 0, "会计报表": 0, "会计账簿": 0, "报表-20": 0}
    # 重新统计（从输出sheet读行数）
    for sn in sheet_stats:
        if sn in [s.title for s in wb_out.worksheets]:
            ws_tmp = wb_out[sn]
            sheet_stats[sn] = ws_tmp.max_row - 2   # 减去说明行和表头行

    for sn, cnt in sheet_stats.items():
        ws_sum.append([sn, cnt, "原始Excel中件号重复，导入时已跳过（保留首条）"])
    ws_sum.append(["合计", total_skipped, ""])

    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 45

    wb_out.save(OUT_PATH)
    print(f"\n✅ 导出完成：{OUT_PATH}")
    print(f"   合计跳过记录：{total_skipped} 条")

if __name__ == "__main__":
    main()
