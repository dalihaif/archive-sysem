"""
会计档案专用导入脚本
- 支持4个Sheet：会计凭证、会计报表、会计账簿、报表-20
- category统一为"会计"
- archive_number = Sheet名 + "-" + 件号（用于去重）
- 增量导入，按(archive_number, category)去重
- 字段映射：分类→class_number, 件号→file_number, 文件题名→title,
           责任者→responsible, 文件编号→doc_number, 形成时间→doc_date,
           保管期限→retention_period, 页数→pages, 主题词→keywords,
           附注→remarks, 存放位置→electronic_location
"""

import sys
import os
import datetime
import hashlib

sys.path.insert(0, os.path.dirname(__file__))

EXCEL_PATH = r"E:\工作文档\档案室\02_归档文件目录与总库\档案总库\会计档案.xlsx"

# 导入的会计档案 Sheet -> 说明
SHEETS = ["会计凭证", "会计报表", "会计账簿", "报表-20"]

# 会计档案字段映射（Excel列名 -> Archive字段）
KUAIJI_FIELD_MAP = {
    "分类":     "class_number",
    "件号":     "file_number",
    "文件题名": "title",
    "责任者":   "responsible",
    "文件编号": "doc_number",
    "形成时间": "doc_date",
    "保管期限": "retention_period",
    "页数":     "pages",
    "主题词":   "keywords",
    "附注":     "remarks",
    "存放位置": "electronic_location",
}

CATEGORY = "会计"
BATCH_NAME = "kuaiji_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


def parse_date(value):
    """解析日期字段"""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    if not s or s.lower() in ("none", "null", "nan"):
        return None
    import re
    patterns = [
        (r"^(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})[日]?$",),
        (r"^(\d{4})(\d{2})(\d{2})$",),
    ]
    for (pattern,) in patterns:
        m = re.match(pattern, s)
        if m:
            g = m.groups()
            try:
                return datetime.date(int(g[0]), int(g[1]), int(g[2]))
            except Exception:
                pass
    # 只有年份
    m = re.match(r"^(\d{4})$", s)
    if m:
        try:
            return datetime.date(int(m.group(1)), 1, 1)
        except Exception:
            pass
    return None


def compute_hashes(data: dict) -> tuple:
    content = ";".join(f"{k}={v}" for k, v in sorted(data.items()) if v)
    return (
        hashlib.sha256(content.encode("utf-8")).hexdigest(),
        hashlib.md5(content.encode("utf-8")).hexdigest(),
    )


def parse_sheet(ws, sheet_name: str) -> list[dict]:
    """解析单个Sheet，返回字典列表"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue

        data = {}
        for i, header in enumerate(headers):
            if header not in KUAIJI_FIELD_MAP:
                continue
            field = KUAIJI_FIELD_MAP[header]
            value = row[i] if i < len(row) else None
            if value is None:
                continue

            if field == "file_number":
                try:
                    v = str(value).strip()
                    data[field] = int(float(v))
                except Exception:
                    # 件号可能是"     1"这种带空格的字符串
                    try:
                        data[field] = int(str(value).strip())
                    except Exception:
                        data[field] = None
            elif field == "doc_date":
                data[field] = parse_date(value)
            else:
                val_str = str(value).strip()
                if val_str and val_str.lower() not in ("none", "null", "nan"):
                    data[field] = val_str

        title = data.get("title", "")
        file_num = data.get("file_number")

        if not title and file_num is None:
            continue  # 空行跳过

        # 构建 archive_number = Sheet名 + "-" + 件号
        archive_number = f"{sheet_name}-{file_num}" if file_num is not None else f"{sheet_name}-{title[:20]}"
        data["archive_number"] = archive_number

        # 尝试从class_number中提取归档年度
        class_num = data.get("class_number", "")
        if class_num:
            import re
            m = re.match(r"^(\d{4})", class_num)
            if m:
                try:
                    data["archive_year"] = int(m.group(1))
                except Exception:
                    pass

        bc_hash, bc_md5 = compute_hashes(data)
        data["bc_hash"] = bc_hash
        data["bc_md5"] = bc_md5

        records.append(data)

    return records


def main():
    import openpyxl
    from app import create_app
    from app.models import Archive
    from app.extensions import db

    app = create_app()

    print(f"\n{'='*60}")
    print(f"  会计档案导入脚本  v1.0")
    print(f"  文件：{EXCEL_PATH}")
    print(f"  批次：{BATCH_NAME}")
    print(f"{'='*60}\n")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    all_records = []  # (sheet_name, data)

    for sn in SHEETS:
        if sn not in wb.sheetnames:
            print(f"[警告] Sheet [{sn}] 不存在，跳过")
            continue
        ws = wb[sn]
        recs = parse_sheet(ws, sn)
        print(f"  Sheet [{sn}]: 解析到 {len(recs)} 条有效记录")
        for r in recs:
            all_records.append(r)

    wb.close()
    print(f"\n  合计解析：{len(all_records)} 条\n")

    with app.app_context():
        # 加载已有会计档号（高效去重）
        existing_keys = set()
        rows = db.session.query(Archive.archive_number, Archive.category).filter(
            Archive.category == CATEGORY
        ).all()
        for an, cat in rows:
            existing_keys.add((an, cat))
        print(f"  数据库中已有会计档案：{len(existing_keys)} 条")

        imported = 0
        skipped = 0
        batch = []

        for data in all_records:
            archive_number = data.get("archive_number", "")
            key = (archive_number, CATEGORY)
            if key in existing_keys:
                skipped += 1
                continue

            archive = Archive(
                category=CATEGORY,
                archive_number=archive_number,
                fonds_number="",
                archive_year=data.get("archive_year"),
                retention_period=data.get("retention_period", ""),
                class_number=data.get("class_number", ""),
                file_number=data.get("file_number"),
                title=data.get("title", ""),
                responsible=data.get("responsible", ""),
                doc_number=data.get("doc_number", ""),
                doc_date=data.get("doc_date"),
                pages=str(data.get("pages", "")) if data.get("pages") else "",
                keywords=data.get("keywords", ""),
                remarks=data.get("remarks", ""),
                electronic_location=data.get("electronic_location", ""),
                bc_hash=data.get("bc_hash", ""),
                bc_md5=data.get("bc_md5", ""),
                import_batch=BATCH_NAME,
            )
            batch.append(archive)
            imported += 1
            existing_keys.add(key)

            if len(batch) >= 1000:
                db.session.add_all(batch)
                db.session.commit()
                print(f"  已提交 {imported} 条...")
                batch = []

        if batch:
            db.session.add_all(batch)
            db.session.commit()

    print(f"\n{'='*60}")
    print(f"  导入完成！")
    print(f"  ✅ 成功导入：{imported} 条")
    print(f"  ⏭️  跳过（已存在）：{skipped} 条")
    print(f"  📦 批次号：{BATCH_NAME}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
